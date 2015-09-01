import json as json_module
import csv
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from core.helpers import clean_string
import re, unicodedata

class Emitter:

    """ DATAL speaks different formats """

    def __init__(self, json):
        self.loaded_json = json
        self.parsed_json = None
        self.result = None
        self.is_array = False
        self.length = 0
        self._parse_junar_json()

    def clean_string(self, p_string):
        p_string = unicodedata.normalize('NFKD', p_string).encode('ascii', 'ignore')
        p_string = re.sub('[^a-zA-Z0-9]+', '-', p_string.strip())
        p_string = re.sub('\-+', '-', p_string)
        p_string = re.sub('\-$', '', p_string)
        return p_string

    def _parse_junar_json(self):

        if self.loaded_json['fType'] != 'ARRAY':

            if self.loaded_json['fType'] != 'NUMBER':
                self.parsed_json = self.loaded_json['fStr'].encode('utf-8', 'ignore')
            else:
                self.parsed_json = self.loaded_json['fNum']
        else:
            self.parsed_json = []
            self.is_array = True
            l_i = 0
            for l_row_number in range(0,self.loaded_json['fRows']):
                l_row  = []
                for l_column_number in range(0, self.loaded_json['fCols']):
                    l_cell = self.loaded_json['fArray'][l_i]
                    if l_cell['fType'] != 'NUMBER':
                        l_row.append(l_cell['fStr'].encode('utf-8', 'ignore'))
                    else:
                        l_row.append(l_cell['fNum'])
                    l_i = l_i + 1
                self.parsed_json.append(l_row)

        if self.loaded_json.has_key('fLength'):
            self.length = self.loaded_json['fLength']


class JSONEmitter(Emitter):
    """ Native Javascript / JSON array of arrays or number or string """

    def __init__(self, json):
        Emitter.__init__(self, json)

    def render(self):
        self.result = json_module.dumps(self.parsed_json)
        return self.result

class CSVEmitter(Emitter):

    """ CSV emitter """

    def __init__(self, json, name):
        self.delimiter  = ','
        self.quotechar  = '"'
        self.csv_writer = None
        self.dialect    = csv.excel
        self.filename   = self.clean_string(name).replace(' ', '_')
        self.headers    = {'Content-Type': 'text/csv;charset=utf-8'
                           , 'Content-Disposition': 'attachment; filename=%s.csv' % self.filename
                          }
        Emitter.__init__(self, json)

    def render(self):

        self.result = CSVFile()
        self.csv_writer = csv.writer(self.result, delimiter=self.delimiter, quotechar=self.quotechar, quoting=csv.QUOTE_ALL, dialect=self.dialect)

        if self.is_array:
            for l_row in self.parsed_json:
                self.csv_writer.writerow(l_row)
        else:
            self.csv_writer.writerow([self.parsed_json])

        return self.result.content

    def get_headers(self):
        return self.headers

class TSVEmitter(CSVEmitter):

    """ TSV emitter """

    def __init__(self, json, name):
        CSVEmitter.__init__(self, json, name)
        self.delimiter = '\t'
        self.quotechar = '"'
        self.dialect = csv.excel_tab
        self.headers    = {'Content-Type': 'text/tab-separated-values;charset=utf-8'
                           , 'Content-Disposition': 'attachment; filename=%s.csv' % self.filename
                          }

class ExcelEmitter(Emitter):

    """ Excel emitter """

    def __init__(self, json, name):
        self.workbook = None
        self.name     = self.clean_string(name)
        self.filename = self.name.replace(' ', '_')
        self.headers  = {'Content-Type': 'application/vnd.ms-excel;charset=utf-8'
                         , 'Content-Disposition': 'attachment; filename=%s.xlsx' % self.filename
                        }
        Emitter.__init__(self, json)

    def render(self):

        self.workbook = Workbook(encoding = 'utf-8')
        worksheet0 = self.workbook.get_active_sheet()
        worksheet0.title = self.name[:31]

        if self.is_array:
            l_row_number = 0
            for l_row in self.parsed_json:
                l_column_number = 0
                for l_cell in l_row:
                    cell = worksheet0.cell(row = l_row_number, column = l_column_number)
                    cell.value = l_cell
                    l_column_number = l_column_number + 1
                l_row_number = l_row_number + 1
        else:
            cell = worksheet0.cell(row = 0, column = 0)
            cell.value = self.parsed_json

        return save_virtual_workbook(self.workbook)

    def get_headers(self):
        return self.headers

class HTMLEmitter(Emitter):

    """ HTML emitter """

    def __init__(self, json, name):
        self.name = name
        self.headers = {'Content-Type': 'text/html;charset=utf-8'}
        Emitter.__init__(self, json)

    def _parse_junar_json(self):
        # we dont need to do anything here
        pass

    def render(self):
        caption_html = '<caption>%s</caption>' % self.name
        table_wrapper_template_open  = '<table itemscope itemtype="http://schema.org/Table">' + caption_html
        table_wrapper_template_close = '</table>'

        if self.loaded_json['fType'] != 'ARRAY':
            cell_html = '<tr>' + self._get_cell_html(self.loaded_json) + '</tr>'
            return table_wrapper_template_open + cell_html + table_wrapper_template_close
        else:
            html = ''
            i = 0
            for row in range(0,self.loaded_json['fRows']):
                table_row = '<tr>%s</tr>'
                cells_html  = ''
                for columns in range(0, self.loaded_json['fCols']):
                    cell  = self.loaded_json['fArray'][i]
                    cells_html += self._get_cell_html(cell)
                    i += 1
                html += table_row % cells_html
            return table_wrapper_template_open + html + table_wrapper_template_close

    def get_headers(self):
        return self.headers

    def _get_cell_html(self, cell):

        if cell['fType'] == 'TEXT':
            schema_url =  'http://schema.org/Text'
            value = cell['fStr']

        elif cell['fType'] == 'NUMBER':
            if type(cell['fNum']).__name__=='int':
                schema_url = 'http://schema.org/Integer'
            else:
                schema_url = 'http://schema.org/Float'
            value = str(cell['fNum'])

        elif cell['fType'] == 'LINK':
            link_template = '<a target="_blank" href="%s" rel="nofollow" title="%s">%s</a>'
            schema_url = 'http://schema.org/URL'
            value = link_template % (cell['fUri'], cell['fStr'], cell['fStr'])

        elif cell['fType'] == 'ERROR':
            return '<td>No data found</td>'

        return '<td itemscope itemtype="%s">%s</td>' % (schema_url, value)


class XMLEmitter:

    """ Xml emitter """

    def __init__(self, xml, name = ''):
        self.xml = xml
        self.headers    = {'Content-Type': 'text/xml;charset=utf-8'}

    def get_headers(self):
        return self.headers

    def render(self):
        return self.xml

class CSVFile:

    def __init__(self):
        self.content = ''

    def write(self, p_content):
        self.content = self.content + p_content

    def __str__(self):
       return self.content

    def __repr__(self):
       return self.content
